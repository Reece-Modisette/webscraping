from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font, Alignment, numbers

import keys
from twilio.rest import Client
client = Client(keys.accountSID, keys.auth_token)

TWnum = '+1 507 570 5273'
myphone = "+19366356523"

#Open the Webpage and test connection
webpage = 'https://www.tradingview.com/markets/cryptocurrencies/prices-all/'
page = urlopen(webpage)			
req = Request(webpage)

webpage = urlopen(req).read()

soup = BeautifulSoup(page, 'html.parser')

print()
print(soup.title.text)



#Excel Workbook Creation
wb = xl.Workbook()
ws = wb.active
ws.title = 'Web-Scraping Project'

ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 17

ws['A1'] = 'No.'
ws['B1'] = 'Crypto Name'
ws['C1'] = 'Symbol'
ws['D1'] = 'Current Price'
ws['E1'] = '24Hrs Previous Price'
ws['F1'] = '% Change Past 24Hrs'

wrap_style = Alignment(wrap_text=True)

ws['A1'].alignment = wrap_style
ws['B1'].alignment = wrap_style
ws['C1'].alignment = wrap_style
ws['D1'].alignment = wrap_style
ws['E1'].alignment = wrap_style
ws['F1'].alignment = wrap_style

x = 1
#Web-Scraping Loop
table_rows = soup.findAll('tr')

for row in table_rows[1:6]:
    td = row.findAll('td')
    sup = row.findAll('sup')
    a = row.findAll('a')
    img= row.findAll('img')

    no = x
    price_cur = td[2]
    percent_change = td[3]
    name = sup[0].text
    symbol = a[0].text 
    price_cur = float(price_cur.text.replace('USD',''))
    #percentage_change
    if '−' in percent_change.text:
        percent_change = float(percent_change.text.replace('%','').replace('−','-'))
    else:
        percent_change = float(percent_change.text.replace('%',''))
    
    multiplier = percent_change/100
    yesterday_price = round(price_cur/(1+multiplier),6)

    # print()
    # print("Crypto:",name)
    # print("Symbol:",symbol)
    # print("Curr $:",price_cur)
    # print("24hr %:",percent_change)
    # print("Yest $:",yesterday_price)


    #Twilio Text message "if" code
    #----------------------------------------------------------------
    if name == 'Bitcoin'or name == 'USD Coin':
        if abs(yesterday_price - price_cur) >5:
            change = abs(yesterday_price - price_cur)
            mytext = name + " price changed by: " +str(change)+". Go check it out while the gettin's still good!"

            textmsg = client.messages.create(to=myphone, from_=TWnum, body=mytext)
            print(textmsg.status)

    #writing data into Excel Workbook
    #----------------------------------------------------------------
    x+=1
    ws['A'+str(x)] = no
    ws['B'+str(x)] = name
    ws['C'+str(x)] = symbol
    ws['D'+str(x)] = price_cur
    ws['E'+str(x)] = yesterday_price
    ws['F'+str(x)] = percent_change

header_font = Font(size=16, bold=True, color='33CC33', name='Rockwell')
body_font = Font(size=12, color='FFC000', name='Times New Roman')

for cell in ws[1:1]:
    cell.font = header_font

cell_range = ws['A2:F6']
for row in cell_range:
    for cell in row:
        cell.font = body_font

cell_range = ws['F2:F6']
for row in cell_range:
    for cell in row:
        cell.number_format = numbers.FORMAT_PERCENTAGE_00

cell_range = ws['D2:E6']
for row in cell_range:
    for cell in row:
        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


wb.save('Web-ScrapingProject.xlsx')
